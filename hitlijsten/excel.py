"""Excel-bouwer: per hitlijst twee werkboeken uit de database.

Voor lijst "top40" (bestand="Top40", jaar 2026):

* ``Top40_2026.xlsx``      -- tab per week met de COMPLETE lijst van die week,
  waarin de nieuwe binnenkomers een lichtblauwe rij krijgen, plus een tab
  "Totaal" met het puntenklassement van het hele jaar.
* ``Top40_Jaar_2026.xlsx`` -- tab "Jaaroverzicht": matrix nummer x week met de
  positie in elke cel.

De database is de enige bron. Er wordt nooit iets opgehaald vanuit deze module.

Twee begrippen die makkelijk door elkaar lopen:

"nieuw"       (onze definitie)  de sleutel kwam dit jaar niet eerder voor in
                                DEZE lijst. Op de eerste beschikbare week is dus
                                de hele lijst nieuw; daarna alleen echte
                                nieuwkomers. Een re-entry is niet nieuw.
"site-status" (van de site)     wat de site zelf zegt: nieuw/stijger/daler/
                                gelijk/terug/onbekend. Daaraan zie je of iets
                                een echte binnenkomer is of een nummer dat al
                                liep toen wij begonnen met verzamelen.
"""
from __future__ import annotations

import sqlite3
from contextlib import nullcontext
from datetime import date
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import LIJSTEN, excel_map
from .datums import vrijdag_van
from .db import (
    Looptijd, alle_jaren, looptijden, noteringen_van_jaar, totalen_over,
    verbinding,
)

__all__ = [
    "BestandInGebruik",
    "Nummer",
    "LijstGegevens",
    "bouw_alles",
    "bouw_decennium",
    "bouw_decennium_werkboek",
    "bouw_totalen_werkboek",
    "bouw_lijst",
    "verzamel_lijst",
    "bewaar_werkboek",
    "tabnaam",
]


class BestandInGebruik(Exception):
    """Het doelbestand kon niet worden geschreven; meestal staat het open."""


# --- opmaak ----------------------------------------------------------------

KOP_FONT = Font(bold=True)
TOELICHTING_FONT = Font(italic=True, color="FF808080")
# Lagere positie = beter; alleen de top 3 krijgt een accent, oplopend flauwer.
TOP_KLEUREN = {
    1: PatternFill("solid", fgColor="FFFFD966"),
    2: PatternFill("solid", fgColor="FFFFE699"),
    3: PatternFill("solid", fgColor="FFFFF2CC"),
}
# Lichtblauw over de hele rij voor nummers die dit jaar nieuw binnenkomen.
# Excel-standaardtint "Blauw, accent 1, lichter 80%" -- valt op zonder de tekst
# onleesbaar te maken, ook bij afdrukken in grijstinten.
NIEUW_VULLING = PatternFill("solid", fgColor="FFDDEBF7")

# Korte, getalachtige kolommen lezen prettiger gecentreerd; de kopregel blijft
# links uitgelijnd zodat het filterpijltje op zijn plek blijft.
GECENTREERD = ("Positie", "Vorige positie", "Aantal weken", "Site-status")
GECENTREERD_TOTAAL = (
    "Punten",
    "Hoogste positie",
    "Aantal weken genoteerd",
    "Weken volgens site",
    "Binnenkomst",
    "Laatste notering",
    "Alarmschijf",
    "Dancesmash",
)

MIN_BREEDTE = 6
MAX_BREEDTE = 44

# Excel staat deze tekens niet toe in een tabnaam, en niet meer dan 31 tekens.
VERBODEN_TABTEKENS = ':\\/?*[]'

WEEK_TOELICHTING = (
    "De complete lijst van deze week. Lichtblauw = nummers die dit jaar niet "
    "eerder in deze lijst stonden (een re-entry telt dus niet als nieuw). De "
    "kolommen 'Site-status' en 'Aantal weken' komen van de site zelf: "
    "site-status 'nieuw' is een echte binnenkomer, 'terug' een re-entry, de "
    "rest liep al toen wij begonnen."
)


def tabnaam(naam: str) -> str:
    """Maak een tekst bruikbaar als tabnaam (max 31 tekens, geen : \\ / ? * [ ])."""
    schoon = "".join(" " if teken in VERBODEN_TABTEKENS else teken for teken in naam)
    schoon = " ".join(schoon.split()).strip("'")
    return (schoon[:31] or "Blad")


# --- gegevens uit de database ----------------------------------------------


@dataclass
class Nummer:
    """Eén uniek nummer (sleutel) binnen één lijst en één jaar."""

    sleutel: str
    titel: str
    artiest: str
    label: Optional[str] = None
    # week -> beste (laagste) positie in die week
    posities: dict[int, int] = field(default_factory=dict)
    punten: int = 0
    weken_volgens_site: Optional[int] = None

    @property
    def hoogste_positie(self) -> int:
        return min(self.posities.values())

    @property
    def eerste_week(self) -> int:
        return min(self.posities)

    @property
    def laatste_week(self) -> int:
        return max(self.posities)

    @property
    def aantal_weken(self) -> int:
        """Zelf geteld uit onze data, niet het weken_genoteerd-veld van de site."""
        return len(self.posities)


@dataclass
class LijstGegevens:
    """Alles wat de twee werkboeken van één lijst nodig hebben."""

    lijst: str
    naam: str
    heeft_label: bool
    bestand: str
    weken: list[int]
    lengte_per_week: dict[int, int]
    nummers: dict[str, Nummer]
    # week -> rijen (sqlite3.Row) van de nummers die die week voor het eerst
    # in deze lijst stonden, oplopend op positie
    nieuw_per_week: dict[int, list[sqlite3.Row]]
    # week -> alle rijen van die week, oplopend op positie (de complete lijst)
    alle_per_week: dict[int, list[sqlite3.Row]]


def _sleutels_van_vorig_jaar(
    con: sqlite3.Connection, lijst: str, jaar: int
) -> set[str]:
    """Welke nummers noteerden vorig jaar al in deze lijst?

    Alleen bedoeld om week 1 eerlijk te maken: een nummer dat over de
    jaarwisseling heen doorloopt is geen binnenkomer. Hebben we vorig jaar
    helemaal niet verzameld, dan valt de site-status in als noodrem (zie
    `_valt_terug_op_site`), anders zou de eerste week van onze allereerste
    jaargang alsnog volledig blauw kleuren.
    """
    return {
        r["sleutel"]
        for r in con.execute(
            "SELECT DISTINCT sleutel FROM noteringen WHERE lijst=? AND jaar=?",
            (lijst, jaar - 1),
        )
    }


def verzamel_lijst(con: sqlite3.Connection, lijst: str, jaar: int) -> Optional[LijstGegevens]:
    """Lees één lijst uit de database. Geeft None als er geen enkele rij is."""
    rijen = noteringen_van_jaar(con, lijst, jaar)
    if not rijen:
        return None

    cfg = LIJSTEN.get(lijst, {})

    # Lengte van de lijst per week = hoogste positienummer dat die week in de
    # database staat. Per week bepaald, want de Tipparade varieert.
    lengte_per_week: dict[int, int] = {}
    # Per week per sleutel de beste rij. Een sleutel hoort maar één keer per week
    # voor te komen; staat hij er toch twee keer (twee versies van hetzelfde
    # nummer), dan telt de hoogste notering en wordt de week één keer geteld.
    per_week: dict[int, dict[str, sqlite3.Row]] = {}
    for rij in rijen:
        week = rij["week"]
        lengte_per_week[week] = max(lengte_per_week.get(week, 0), rij["positie"])
        vak = per_week.setdefault(week, {})
        bestaand = vak.get(rij["sleutel"])
        if bestaand is None or rij["positie"] < bestaand["positie"]:
            vak[rij["sleutel"]] = rij

    weken = sorted(per_week)
    nummers: dict[str, Nummer] = {}
    nieuw_per_week: dict[int, list[sqlite3.Row]] = {}
    alle_per_week: dict[int, list[sqlite3.Row]] = {}

    # Wat liep er al aan het eind van vorig jaar? Zonder dat zou in week 1 de
    # hele lijst als binnenkomer gelden, terwijl de meeste nummers gewoon
    # doorlopen over de jaarwisseling heen.
    liep_al = _sleutels_van_vorig_jaar(con, lijst, jaar)

    # Is er van vorig jaar niets, dan is dit onze oudste jaargang. Voor de
    # eerste week vertrouwen we dan op wat de site zelf zegt: alleen een echte
    # binnenkomer of re-entry telt als nieuw, de rest liep al.
    if not liep_al and weken:
        liep_al = {
            sleutel
            for sleutel, rij in per_week[weken[0]].items()
            if rij["site_status"] not in ("nieuw", "terug")
        }

    for week in weken:
        nieuw: list[sqlite3.Row] = []
        for sleutel, rij in per_week[week].items():
            if sleutel in liep_al:
                # Al bekend uit vorig jaar: wel meetellen, niet als nieuw
                # markeren. De eerste toekenning hieronder maakt het nummer aan.
                liep_al.discard(sleutel)
                nummers.setdefault(
                    sleutel,
                    Nummer(sleutel=sleutel, titel=rij["titel"], artiest=rij["artiest"]),
                )
            nummer = nummers.get(sleutel)
            if nummer is None:
                nummer = Nummer(sleutel=sleutel, titel=rij["titel"], artiest=rij["artiest"])
                nummers[sleutel] = nummer
                nieuw.append(rij)
            # Weken lopen oplopend, dus de laatste toekenning is de meest
            # recente schrijfwijze. Een leeg label overschrijft een eerder
            # gevonden label niet: dat is eerder een gat dan een naamswijziging.
            nummer.titel = rij["titel"]
            nummer.artiest = rij["artiest"]
            if rij["label"]:
                nummer.label = rij["label"]
            nummer.posities[week] = rij["positie"]
            nummer.punten += lengte_per_week[week] - rij["positie"] + 1
            volgens_site = rij["weken_genoteerd"]
            if volgens_site is not None:
                nummer.weken_volgens_site = max(nummer.weken_volgens_site or 0, volgens_site)
        nieuw_per_week[week] = sorted(nieuw, key=lambda r: r["positie"])
        alle_per_week[week] = sorted(per_week[week].values(), key=lambda r: r["positie"])

    return LijstGegevens(
        lijst=lijst,
        naam=cfg.get("naam", lijst),
        heeft_label=bool(cfg.get("heeft_label", False)),
        bestand=cfg.get("bestand", lijst),
        weken=weken,
        lengte_per_week=lengte_per_week,
        nummers=nummers,
        nieuw_per_week=nieuw_per_week,
        alle_per_week=alle_per_week,
    )


# --- schrijven -------------------------------------------------------------


def _schrijf_tabel(
    ws: Worksheet,
    kolommen: Sequence[str],
    rijen: Iterable[Sequence[Any]],
    *,
    toelichting: Optional[str] = None,
    accent: Optional[Sequence[bool]] = None,
    centreer: Optional[Sequence[str]] = None,
) -> int:
    """Zet kop + gegevens neer met filter, bevroren kop en kolombreedtes.

    `accent` loopt gelijk op met `rijen`: waar True staat krijgt de hele rij een
    achtergrondkleur over alle kolommen.

    Geeft het rijnummer van de kopregel terug.
    """
    kop_rij = 1
    if toelichting:
        cel = ws.cell(row=1, column=1, value=toelichting)
        cel.font = TOELICHTING_FONT
        cel.alignment = Alignment(vertical="center")
        kop_rij = 2

    breedtes = [len(str(kop)) for kop in kolommen]
    for kolom, kop in enumerate(kolommen, start=1):
        cel = ws.cell(row=kop_rij, column=kolom, value=kop)
        cel.font = KOP_FONT

    # Kolommen die gecentreerd worden -- alleen de gegevens, niet de kop.
    te_centreren = {
        kolom for kolom, kop in enumerate(kolommen, start=1)
        if centreer and kop in centreer
    }

    aantal = 0
    for aantal, rij in enumerate(rijen, start=1):
        gemarkeerd = bool(accent and aantal <= len(accent) and accent[aantal - 1])
        for kolom, waarde in enumerate(rij, start=1):
            cel = ws.cell(row=kop_rij + aantal, column=kolom, value=waarde)
            if kolom in te_centreren:
                cel.alignment = Alignment(horizontal="center")
            # Als echte datum wegschrijven, niet als tekst: zo kun je er in
            # Excel op sorteren en filteren en ermee rekenen.
            if isinstance(waarde, date):
                cel.number_format = "DD/MM/YYYY"
            # Over alle kolommen, ook de lege: anders houdt de markering
            # halverwege de rij op.
            if gemarkeerd:
                cel.fill = NIEUW_VULLING
            if waarde is not None:
                breedtes[kolom - 1] = max(breedtes[kolom - 1], len(str(waarde)))

    laatste_kolom = get_column_letter(len(kolommen))
    ws.freeze_panes = f"A{kop_rij + 1}"
    ws.auto_filter.ref = f"A{kop_rij}:{laatste_kolom}{kop_rij + aantal}"
    for kolom, breedte in enumerate(breedtes, start=1):
        ws.column_dimensions[get_column_letter(kolom)].width = min(
            MAX_BREEDTE, max(MIN_BREEDTE, breedte + 2)
        )
    return kop_rij


def _weektab(wb: Workbook, gegevens: LijstGegevens, week: int) -> None:
    """De complete lijst van die week, met de binnenkomers lichtblauw."""
    kolommen = ["Positie", "Vorige positie", "Artiest", "Titel"]
    if gegevens.heeft_label:
        kolommen.append("Label")
    kolommen += ["Aantal weken", "Site-status", "Sleutel"]

    nieuwe_sleutels = {r["sleutel"] for r in gegevens.nieuw_per_week[week]}

    rijen: list[list[Any]] = []
    accent: list[bool] = []
    for rij in gegevens.alle_per_week[week]:
        vorige = rij["vorige_positie"]
        waarden: list[Any] = [
            int(rij["positie"]),
            int(vorige) if vorige is not None else None,
            rij["artiest"],
            rij["titel"],
        ]
        if gegevens.heeft_label:
            waarden.append(rij["label"])
        weken_site = rij["weken_genoteerd"]
        waarden += [
            int(weken_site) if weken_site is not None else None,
            rij["site_status"],
            rij["sleutel"],
        ]
        rijen.append(waarden)
        accent.append(rij["sleutel"] in nieuwe_sleutels)

    toelichting = WEEK_TOELICHTING
    if not any(accent):
        toelichting = "Geen nieuwe nummers deze week. " + toelichting

    ws = wb.create_sheet(tabnaam(f"Week {week:02d}"))
    _schrijf_tabel(ws, kolommen, rijen, toelichting=toelichting, accent=accent,
                   centreer=GECENTREERD)


def _grensmarkering(loop: Looptijd) -> Optional[str]:
    """Korte aanduiding waarom een datum buiten dit jaar valt."""
    if loop.begon_eerder and loop.loopt_door:
        return "begon eerder en loopt door"
    if loop.begon_eerder:
        return "begon vorig jaar"
    if loop.loopt_door:
        return "loopt door"
    return None


def _totaaltab(wb: Workbook, gegevens: LijstGegevens, jaar: int, con=None) -> None:
    # Alarmschijven en Dancesmashes komen van michajans.nl en bestaan alleen
    # voor de Top 40. Ontbreken ze in de database, dan blijven de kolommen weg
    # in plaats van leeg -- lege kolommen suggereren dat er niets toegekend is.
    onderscheiding = {}
    correcties: dict[str, dict] = {}
    if gegevens.lijst == "top40":
        try:
            from .onderscheidingen import per_sleutel

            onderscheiding = per_sleutel(jaar, con)
        except Exception:
            onderscheiding = {}
        try:
            from .kruiscontrole import correcties_voor

            correcties = correcties_voor(jaar, gegevens.lijst, con)
        except Exception:
            correcties = {}

    kolommen = ["Artiest", "Titel"]
    if gegevens.heeft_label:
        kolommen.append("Label")
    kolommen += [
        "Punten",
        "Hoogste positie",
        "Aantal weken genoteerd",
        "Weken volgens site",
        "Binnenkomst",
        "Laatste notering",
        "Loopt over jaargrens",
    ]
    looptijd = looptijden(con, gegevens.lijst, jaar) if con is not None else {}
    if onderscheiding:
        kolommen += ["Alarmschijf", "Dancesmash"]
    if correcties:
        kolommen.append("Bron")
    kolommen.append("Sleutel")

    def cijfers(nummer: Nummer) -> tuple[int, int, int, Optional[str]]:
        """Punten, hoogste positie en weken -- gecorrigeerd waar dat vastligt.

        Bij een groot verschil met michajans.nl geldt zijn cijfer: hij haalt de
        fouten uit de officiele lijst. De weektabs en de jaarmatrix blijven wel
        onze eigen waarneming, want zijn jaarlijst geeft geen posities per week.
        Voor zo'n rij is het jaartotaal dus niet de som van de weektabs; de
        kolom Bron zegt waar het vandaan komt.
        """
        c = correcties.get(nummer.sleutel)
        if c:
            return c["punten"], c["hoogste"], c["weken"], c["bron"]
        return nummer.punten, nummer.hoogste_positie, nummer.aantal_weken, None

    gesorteerd = sorted(
        gegevens.nummers.values(),
        key=lambda n: (-cijfers(n)[0], cijfers(n)[1], n.eerste_week, n.titel.lower()),
    )
    rijen = []
    for nummer in gesorteerd:
        punten, hoogste, weken, bron = cijfers(nummer)
        waarden: list[Any] = [nummer.artiest, nummer.titel]
        if gegevens.heeft_label:
            waarden.append(nummer.label)
        waarden += [
            int(punten),
            int(hoogste),
            int(weken),
            int(nummer.weken_volgens_site) if nummer.weken_volgens_site is not None else None,
        ]
        # De uitzenddatum in plaats van het weeknummer. Twee dingen kunnen dan
        # buiten dit jaar vallen, en allebei is het de echte datum en geen fout:
        # bij negen jaargangen ligt de vrijdag van week 1 op 31 december van het
        # jaar ervoor, en een notering die over de jaarwisseling doorloopt begint
        # of eindigt gewoon in het buurjaar.
        loop = looptijd.get(nummer.sleutel)
        if loop is None:
            waarden += [
                vrijdag_van(jaar, nummer.eerste_week),
                vrijdag_van(jaar, nummer.laatste_week),
                None,
            ]
        else:
            waarden += [loop.begin, loop.eind, _grensmarkering(loop)]
        if onderscheiding:
            gekregen = onderscheiding.get(nummer.sleutel, {})
            waarden += [gekregen.get("alarmschijf"), gekregen.get("dancesmash")]
        if correcties:
            waarden.append(bron)
        waarden.append(nummer.sleutel)
        rijen.append(waarden)

    ws = wb.create_sheet(tabnaam("Totaal"))
    _schrijf_tabel(ws, kolommen, rijen, centreer=GECENTREERD_TOTAAL)


def _jaartab(wb: Workbook, gegevens: LijstGegevens) -> None:
    vaste_koppen = ["Artiest", "Titel"] + (["Label"] if gegevens.heeft_label else [])
    kolommen = list(vaste_koppen)
    kolommen += [str(week) for week in gegevens.weken]
    kolommen += ["Hoogste positie", "Aantal weken", "Punten", "Sleutel"]

    gesorteerd = sorted(
        gegevens.nummers.values(),
        key=lambda n: (n.hoogste_positie, n.eerste_week, n.titel.lower()),
    )
    rijen = []
    for nummer in gesorteerd:
        waarden: list[Any] = [nummer.artiest, nummer.titel]
        if gegevens.heeft_label:
            waarden.append(nummer.label)
        for week in gegevens.weken:
            positie = nummer.posities.get(week)
            waarden.append(int(positie) if positie is not None else None)
        waarden += [
            int(nummer.hoogste_positie),
            int(nummer.aantal_weken),
            int(nummer.punten),
            nummer.sleutel,
        ]
        rijen.append(waarden)

    ws = wb.create_sheet(tabnaam("Jaaroverzicht"))
    kop_rij = _schrijf_tabel(ws, kolommen, rijen)

    eerste_weekkolom = len(vaste_koppen) + 1
    for nummer_rij in range(len(rijen)):
        for verschuiving in range(len(gegevens.weken)):
            cel = ws.cell(row=kop_rij + 1 + nummer_rij, column=eerste_weekkolom + verschuiving)
            vulling = TOP_KLEUREN.get(cel.value)
            if vulling is not None:
                cel.fill = vulling


GECENTREERD_DECENNIUM = (
    "Punten", "Hoogste positie", "Aantal weken genoteerd", "Jaargangen",
    "Binnenkomst", "Laatste notering", "Loopt buiten de periode",
)


def bouw_decennium_werkboek(
    con: sqlite3.Connection, lijst: str, decennium: int
) -> Optional[Workbook]:
    """Eén tab met het puntenklassement over tien jaargangen."""
    return bouw_totalen_werkboek(con, lijst, decennium, decennium + 9)


def bouw_totalen_werkboek(
    con: sqlite3.Connection, lijst: str, van: int, tot: int,
    top: Optional[int] = None,
) -> Optional[Workbook]:
    """Eén tab met het puntenklassement over de jaargangen `van` t/m `tot`.

    Alleen zinvol voor een lijst die al die jaren even lang was -- zie
    `totalen_over` en de LEESMIJ. Geeft None als er geen data is. Met `top`
    alleen de bovenste zoveel -- wat er op het scherm gekozen is, hoort er
    ook uit te komen.
    """
    nummers = totalen_over(con, lijst, van, tot)
    if not nummers:
        return None
    if top:
        nummers = nummers[:top]

    cfg = LIJSTEN.get(lijst, {})
    heeft_label = bool(cfg.get("heeft_label"))
    gecorrigeerd = any(n["gecorrigeerd"] for n in nummers)
    # Bij de volledige historie valt er per definitie niets buiten de periode;
    # een kolom die dan overal leeg is suggereert dat er niets te melden viel.
    grensgevallen = any(n["begon_eerder"] or n["loopt_door"] for n in nummers)

    kolommen = ["Artiest", "Titel"]
    if heeft_label:
        kolommen.append("Label")
    kolommen += [
        "Punten",
        "Hoogste positie",
        "Aantal weken genoteerd",
        "Jaargangen",
        "Binnenkomst",
        "Laatste notering",
    ]
    if grensgevallen:
        kolommen.append("Loopt buiten de periode")
    if gecorrigeerd:
        kolommen.append("Bron")
    kolommen.append("Sleutel")

    rijen: list[list[Any]] = []
    for n in nummers:
        waarden: list[Any] = [n["artiest"], n["titel"]]
        if heeft_label:
            waarden.append(n["label"])
        jaren = n["jaren"]
        waarden += [
            n["punten"],
            n["hoogste"],
            n["weken"],
            str(jaren[0]) if len(jaren) == 1 else f"{jaren[0]}-{jaren[-1]}",
            # De datums blijven binnen het decennium; de kolom ernaast zegt of
            # de notering erbuiten doorliep.
            _datum(n["eerste_sorteer"]),
            _datum(n["laatste_sorteer"]),
        ]
        if grensgevallen:
            waarden.append(_buiten_periode(n))
        if gecorrigeerd:
            waarden.append("michajans.nl" if n["gecorrigeerd"] else None)
        waarden.append(n["sleutel"])
        rijen.append(waarden)

    toelichting = (
        (f"De top {top} van het " if top else "Het ") +
        f"puntenklassement {van}-{tot} van de {cfg.get('naam', lijst)}. "
        "Punten per notering = lijstlengte - positie + 1, per jaargang gerekend "
        "en daarna opgeteld: dit klassement is dus de som van de jaarbestanden."
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(tabnaam(f"Klassement {van}-{tot}"))
    _schrijf_tabel(ws, kolommen, rijen, toelichting=toelichting,
                   centreer=GECENTREERD_DECENNIUM)
    return wb


def bouw_week_werkboek(
    con: sqlite3.Connection, lijst: str, jaar: int, week: int
) -> Optional[Workbook]:
    """Eén weeklijst als los werkboek -- dezelfde tab als in het jaarbestand."""
    gegevens = verzamel_lijst(con, lijst, jaar)
    if gegevens is None or week not in gegevens.alle_per_week:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    _weektab(wb, gegevens, week)
    return wb


def bouw_jaarlijksen_werkboek(
    con: sqlite3.Connection, top: Optional[int] = None
) -> Optional[Workbook]:
    """De gecombineerde lijst over alle jaarlijkse lijsten, één tab."""
    from .db import jaarlijkse_totalen

    nummers = jaarlijkse_totalen(con)
    if not nummers:
        return None
    if top:
        nummers = nummers[:top]

    kolommen = ["#", "Artiest", "Titel", "Punten", "Edities", "Lijsten",
                "Hoogste positie", "Hoogste in", "Sleutel"]
    rijen: list[list[Any]] = []
    for nr, n in enumerate(nummers, start=1):
        naam = LIJSTEN.get(n["hoogste_lijst"], {}).get("naam", n["hoogste_lijst"])
        rijen.append([nr, n["artiest"], n["titel"], n["punten"], n["edities"],
                      n["lijsten"], n["hoogste"], f"{naam} {n['hoogste_jaar']}",
                      n["sleutel"]])

    toelichting = (
        "Alle jaarlijkse lijsten samen, genormaliseerd op lijstlengte: elke "
        "notering telt (lengte - positie + 1) / lengte punten, dus de nummer 1 "
        "van elke lijst is precies één punt waard. Het maximum is het aantal "
        "edities."
    )
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(tabnaam("Jaarlijsten totaal"))
    _schrijf_tabel(ws, kolommen, rijen, toelichting=toelichting,
                   centreer=("#", "Punten", "Edities", "Lijsten",
                             "Hoogste positie"))
    return wb


def _datum(iso: str) -> date:
    jaar, maand, dag = (int(deel) for deel in iso.split("-"))
    return date(jaar, maand, dag)


def _buiten_periode(nummer: dict) -> Optional[str]:
    if nummer["begon_eerder"] and nummer["loopt_door"]:
        return "begon eerder en loopt door"
    if nummer["begon_eerder"]:
        return "begon eerder"
    if nummer["loopt_door"]:
        return "loopt door"
    return None


def bouw_decennium(
    con: sqlite3.Connection, lijst: str, decennium: int,
    uitvoer_map: Optional[Path] = None,
) -> list[Path]:
    """Schrijf het decenniumbestand naar de decenniummap."""
    wb = bouw_decennium_werkboek(con, lijst, decennium)
    if wb is None:
        print(f"[excel] {lijst}: geen noteringen in {decennium}'s - overgeslagen")
        return []

    bestand = LIJSTEN.get(lijst, {}).get("bestand", lijst)
    map_ = Path(uitvoer_map) if uitvoer_map is not None else excel_map(decennium).parent
    map_.mkdir(parents=True, exist_ok=True)
    pad = bewaar_werkboek(
        wb, map_ / f"{bestand}_Decennium_{decennium}-{decennium + 9}.xlsx")
    print(f"[excel] {lijst}: {decennium}'s -> {pad.name}")
    return [pad]


def bewaar_werkboek(wb: Workbook, pad: Path) -> Path:
    """Sla op en vertaal een bestandsslot naar een leesbare fout.

    Komt vooral voor als iemand het werkboek via Samba open heeft staan in
    Excel: dan houdt Windows het bestand vast en faalt het schrijven.
    """
    try:
        wb.save(pad)
    except PermissionError as fout:
        raise BestandInGebruik(
            f"Kan {pad} niet opslaan. Het bestand staat waarschijnlijk nog open in "
            f"Excel. Sluit '{pad.name}' en draai opnieuw."
        ) from fout
    return pad


def bouw_lijst(
    con: sqlite3.Connection, lijst: str, jaar: int, uitvoer_map: Optional[Path] = None
) -> list[Path]:
    """Bouw de twee werkboeken van één lijst. Lege lijst -> melding, geen bestand."""
    gegevens = verzamel_lijst(con, lijst, jaar)
    if gegevens is None:
        print(f"[excel] {lijst}: geen noteringen voor {jaar} in de database - overgeslagen")
        return []

    # Pas hierna de map aanmaken: anders blijft er bij een jaargang zonder data
    # een lege map achter.
    map_ = Path(uitvoer_map) if uitvoer_map is not None else excel_map(jaar)

    weekboek = Workbook()
    weekboek.remove(weekboek.active)
    for week in gegevens.weken:
        _weektab(weekboek, gegevens, week)
    _totaaltab(weekboek, gegevens, jaar, con)

    jaarboek = Workbook()
    jaarboek.remove(jaarboek.active)
    _jaartab(jaarboek, gegevens)

    map_.mkdir(parents=True, exist_ok=True)
    paden = [
        bewaar_werkboek(weekboek, map_ / f"{gegevens.bestand}_{jaar}.xlsx"),
        bewaar_werkboek(jaarboek, map_ / f"{gegevens.bestand}_Jaar_{jaar}.xlsx"),
    ]
    print(
        f"[excel] {lijst}: {len(gegevens.weken)} weken, {len(gegevens.nummers)} unieke "
        f"nummers -> {paden[0].name}, {paden[1].name}"
    )
    return paden


def bouw_alles(
    jaar: int,
    *,
    con: Optional[sqlite3.Connection] = None,
    uitvoer_map: Optional[Path] = None,
    lijsten: Optional[Iterable[str]] = None,
) -> list[Path]:
    """Bouw alle werkboeken van `jaar` en geef de geschreven paden terug.

    `con`, `uitvoer_map` en `lijsten` zijn er voor tests en deelruns; normaal
    volstaat ``bouw_alles(2026)``.
    """
    namen = list(lijsten) if lijsten is not None else list(LIJSTEN)
    beheer = nullcontext(con) if con is not None else verbinding()
    paden: list[Path] = []
    with beheer as verbonden:
        for naam in namen:
            paden.extend(bouw_lijst(verbonden, naam, jaar, uitvoer_map))
    return paden


if __name__ == "__main__":  # pragma: no cover
    import sys

    from .config import JAAR

    jaar = int(sys.argv[1]) if len(sys.argv) > 1 else JAAR
    for pad in bouw_alles(jaar):
        print(pad)
