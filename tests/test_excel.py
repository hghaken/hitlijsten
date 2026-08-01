"""Tests voor hitlijsten/excel.py met verzonnen maar realistische data.

Draaien kan met pytest, maar ook zonder:

    C:\\Python313\\python.exe tests\\test_excel.py

De testdata gaat naar een tijdelijke sqlite-database en een tijdelijke map; de
echte data\\hitlijsten.sqlite en de echte Excel-bestanden worden niet aangeraakt.

DE OPZET VAN DE SYNTHETISCHE DATA
---------------------------------
Weken 1, 2, 3, 4, 6 en 7 -- week 5 ontbreekt met opzet (gat in de reeks).

top40 (40 noteringen per week, geen label):
  "Nummer Eén" / De Kampioenen ....... alle zes de weken op #1
  "Terug Van Weggeweest" / Bløf ...... week 1,2,3, weg in week 4, terug in week 6
  "Zomer In De Stad (Radio Edit)" .... binnenkomer in week 3; artiest heet eerst
                                       "Antoon ft. Sef", vanaf week 6
                                       "Antoon feat. Sef" (zelfde sleutel)
  vulling "Liedje 01".."Liedje 41" ... een paar komen en gaan

tipparade: wisselende lengte per week (8, 10, 10, 12, 9, 11)
oranje:    30 noteringen per week mét label; het label van de special
           verandert halverwege het jaar van schrijfwijze
sterrennl: helemaal geen data -- moet overgeslagen worden
"""
from __future__ import annotations

import atexit
import os
import shutil
import sqlite3
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook  # noqa: E402

from hitlijsten import db, excel  # noqa: E402
from hitlijsten.models import Notering  # noqa: E402
from hitlijsten.normalize import sleutel_van  # noqa: E402

JAAR = 2026
WEKEN = [1, 2, 3, 4, 6, 7]  # week 5 ontbreekt met opzet

HIT = ("De Kampioenen", "Nummer Eén")
BLOF = ("Bløf", "Terug Van Weggeweest")
ANTOON_TITEL = "Zomer In De Stad (Radio Edit)"
ANTOON_FT = ("Antoon ft. Sef", ANTOON_TITEL)
ANTOON_FEAT = ("Antoon feat. Sef", ANTOON_TITEL)
TIPHIT = ("Jonge Band", "Nieuwe Tip")
ORANJEHIT = ("Frans Duijts", "Hart Van Mijn Hart")

# Aantal weken dat een nummer volgens de site al noteerde vóór onze eerste week.
VOORSPRONG = {
    HIT: 19,          # site telt straks 20..25 terwijl wij er 6 tellen
    BLOF: 4,
    ANTOON_FT: 0,
    ANTOON_FEAT: 0,
    TIPHIT: 1,
    ORANJEHIT: 7,
}


# --- data opbouwen ---------------------------------------------------------


def _vulling(voorvoegsel: str):
    def maak(index: int) -> tuple[str, str]:
        return (f"{voorvoegsel}Artiest {index:02d}", f"{voorvoegsel}Liedje {index:02d}")

    return maak


def _plaats(vast: dict[int, tuple[str, str]], vulling_idx, lengte: int, maak_vulling):
    """Zet de vaste nummers op hun positie en vul de rest op met vulnummers."""
    vulling_idx = list(vulling_idx)
    vrij = [p for p in range(1, lengte + 1) if p not in vast]
    assert len(vrij) == len(vulling_idx), (
        f"lengte klopt niet: {len(vast)} vast + {len(vulling_idx)} vulling != {lengte}"
    )
    plaatsing = dict(vast)
    for positie, index in zip(vrij, vulling_idx):
        plaatsing[positie] = maak_vulling(index)
    return plaatsing


def _site_status(info, positie, vorige_week, is_eerste_week):
    """Wat de site over deze notering zou zeggen."""
    if info is not None and info["laatste_week"] == vorige_week:
        eerder = info["positie"]
        if eerder == positie:
            return "gelijk", eerder
        return ("stijger" if positie < eerder else "daler"), eerder
    if info is not None:
        return "terug", None  # stond er eerder wel, vorige week niet -> re-entry
    if is_eerste_week:
        # Onze eerste week: wij noemen alles nieuw, de site niet. Elke zevende
        # positie is volgens de site een echte binnenkomer.
        if positie % 7 == 4:
            return "nieuw", None
        return "stijger", positie + 1
    return "nieuw", None


def _schrijf_lijst(con, lijst, weekplannen, label_van=None):
    historie: dict[str, dict] = {}
    vorige_week = None
    for week in sorted(weekplannen):
        plaatsing = weekplannen[week]
        noteringen = []
        for positie in sorted(plaatsing):
            artiest, titel = plaatsing[positie]
            sleutel = sleutel_van(artiest, titel)
            info = historie.get(sleutel)
            status, vorige_positie = _site_status(
                info, positie, vorige_week, vorige_week is None
            )
            if info is None:
                info = historie.setdefault(sleutel, {"aantal": 0})
            info["aantal"] += 1
            info["laatste_week"] = week
            info["positie"] = positie
            noteringen.append(
                Notering(
                    lijst=lijst,
                    jaar=JAAR,
                    week=week,
                    positie=positie,
                    titel=titel,
                    artiest=artiest,
                    label=label_van(artiest, titel, week) if label_van else None,
                    weken_genoteerd=info["aantal"] + VOORSPRONG.get((artiest, titel), 2),
                    vorige_positie=vorige_positie,
                    site_status=status,
                )
            )
        db.bewaar_week(con, lijst, JAAR, week, noteringen)
        vorige_week = week


def _top40_plannen():
    v = _vulling("")
    return {
        1: _plaats({1: HIT, 12: BLOF}, range(1, 39), 40, v),
        2: _plaats({1: HIT, 19: BLOF}, list(range(1, 38)) + [39], 40, v),
        3: _plaats({1: HIT, 31: BLOF, 28: ANTOON_FT}, list(range(1, 37)) + [39], 40, v),
        4: _plaats({1: HIT, 22: ANTOON_FT}, list(range(1, 37)) + [39, 40], 40, v),
        6: _plaats({1: HIT, 38: BLOF, 15: ANTOON_FEAT}, list(range(1, 36)) + [39, 40], 40, v),
        7: _plaats(
            {1: HIT, 35: BLOF, 11: ANTOON_FEAT}, list(range(1, 35)) + [39, 40, 41], 40, v
        ),
    }


def _tipparade_plannen():
    v = _vulling("Tip ")
    lengtes = {1: 8, 2: 10, 3: 10, 4: 12, 6: 9, 7: 11}
    posities = {1: 1, 2: 1, 3: 2, 4: 1, 6: 3, 7: 5}
    return {
        week: _plaats({posities[week]: TIPHIT}, range(1, lengte), lengte, v)
        for week, lengte in lengtes.items()
    }


def _oranje_plannen():
    v = _vulling("Oranje ")
    return {week: _plaats({3: ORANJEHIT}, range(1, 30), 30, v) for week in WEKEN}


def _oranje_label(artiest, titel, week):
    if (artiest, titel) == ORANJEHIT:
        # Zelfde platenmaatschappij, andere schrijfwijze vanaf week 4.
        return "Berk Music" if week <= 3 else "Berk Music BV"
    return "Label " + titel.split()[-1]


# --- eenmalig bouwen -------------------------------------------------------

_gebouwd_cache: dict | None = None


def gebouwd() -> dict:
    """Vul de tijdelijke database en bouw alle werkboeken (eenmalig)."""
    global _gebouwd_cache
    if _gebouwd_cache is not None:
        return _gebouwd_cache

    werkmap = Path(tempfile.mkdtemp(prefix="hitlijsten_test_"))
    atexit.register(shutil.rmtree, werkmap, True)

    con = sqlite3.connect(werkmap / "test.sqlite")
    con.row_factory = sqlite3.Row
    con.executescript(db.SCHEMA)

    _schrijf_lijst(con, "top40", _top40_plannen())
    _schrijf_lijst(con, "tipparade", _tipparade_plannen())
    _schrijf_lijst(con, "oranje", _oranje_plannen(), label_van=_oranje_label)
    # "sterrennl" krijgt met opzet geen enkele rij.
    con.commit()

    uit = werkmap / "uit"
    paden = excel.bouw_alles(JAAR, con=con, uitvoer_map=uit)
    con.close()

    _gebouwd_cache = {"map": uit, "paden": paden}
    return _gebouwd_cache


def _boek(naam: str):
    return load_workbook(gebouwd()["map"] / naam)


def _tabel(ws, kop_rij: int):
    """Lees een tab terug als (koppen, lijst van dicts)."""
    koppen = [cel.value for cel in ws[kop_rij]]
    rijen = []
    for waarden in ws.iter_rows(min_row=kop_rij + 1, values_only=True):
        if all(waarde is None for waarde in waarden):
            continue
        rijen.append(dict(zip(koppen, waarden)))
    return koppen, rijen


def _weektabel(ws):
    return _tabel(ws, 2)  # rij 1 is de toelichting


NIEUW_KLEUR = "FFDDEBF7"   # moet gelijk zijn aan excel.NIEUW_VULLING


def _is_gemarkeerd(ws, rijnummer: int, aantal_kolommen: int) -> bool:
    """Heeft deze rij de lichtblauwe markering, over ALLE kolommen?"""
    kleuren = {
        ws.cell(row=rijnummer, column=k).fill.fgColor.rgb
        for k in range(1, aantal_kolommen + 1)
    }
    if kleuren == {NIEUW_KLEUR}:
        return True
    assert NIEUW_KLEUR not in kleuren, (
        f"rij {rijnummer} is maar half gemarkeerd: {kleuren}"
    )
    return False


def _gemarkeerde_titels(ws, koppen, rijen) -> list[str]:
    """De titels van de rijen die als nieuw gemarkeerd zijn, op volgorde."""
    return [
        rij["Titel"]
        for nr, rij in enumerate(rijen, start=3)   # rij 1 toelichting, 2 kop
        if _is_gemarkeerd(ws, nr, len(koppen))
    ]


# --- tests -----------------------------------------------------------------


def test_bestanden_en_tabbladen():
    resultaat = gebouwd()
    namen = sorted(pad.name for pad in resultaat["paden"])
    assert namen == [
        "OranjeTop30_2026.xlsx",
        "OranjeTop30_Jaar_2026.xlsx",
        "Tipparade_2026.xlsx",
        "Tipparade_Jaar_2026.xlsx",
        "Top40_2026.xlsx",
        "Top40_Jaar_2026.xlsx",
    ], namen
    assert all(pad.exists() for pad in resultaat["paden"])

    # Lijst zonder data: overgeslagen, geen bestand, geen exception.
    assert not (resultaat["map"] / "SterrenNL_2026.xlsx").exists()
    assert not (resultaat["map"] / "SterrenNL_Jaar_2026.xlsx").exists()

    wb = _boek("Top40_2026.xlsx")
    # Week 5 ontbreekt in de database -> geen tab, en Totaal staat achteraan.
    assert wb.sheetnames == [
        "Week 01", "Week 02", "Week 03", "Week 04", "Week 06", "Week 07", "Totaal",
    ], wb.sheetnames
    assert all(len(naam) <= 31 for naam in wb.sheetnames)

    ws = wb["Week 01"]
    assert ws.freeze_panes == "A3"          # toelichting + kop bevroren
    assert ws.auto_filter.ref.startswith("A2:")
    assert ws.cell(row=2, column=1).font.bold
    assert ws.column_dimensions["B"].width > 6

    totaal = wb["Totaal"]
    assert totaal.freeze_panes == "A2"
    assert totaal.auto_filter.ref.startswith("A1:")

    jaar = _boek("Top40_Jaar_2026.xlsx")
    assert jaar.sheetnames == ["Jaaroverzicht"]


def test_weektabs_tonen_de_hele_lijst():
    """Elke weektab is de complete lijst van die week, op positie gesorteerd."""
    wb = _boek("Top40_2026.xlsx")

    koppen, week1 = _weektabel(wb["Week 01"])
    # top40 heeft heeft_label False -> geen Label-kolom.
    assert koppen == [
        "Positie", "Vorige positie", "Artiest", "Titel",
        "Aantal weken", "Site-status", "Sleutel",
    ]
    for naam in ("Week 01", "Week 02", "Week 03", "Week 04", "Week 06", "Week 07"):
        _, rijen = _weektabel(wb[naam])
        assert [r["Positie"] for r in rijen] == list(range(1, 41)), naam

    # Wat de site zegt staat los van onze "nieuw": posities 4,11,18,25,32,39.
    assert sum(1 for r in week1 if r["Site-status"] == "nieuw") == 6

    # Getallen moeten getallen zijn, geen tekst.
    assert isinstance(week1[0]["Positie"], int)
    assert isinstance(week1[0]["Aantal weken"], int)
    # Vorige positie is leeg in onze eerste week bij een echte binnenkomer.
    binnenkomer = next(r for r in week1 if r["Site-status"] == "nieuw")
    assert binnenkomer["Vorige positie"] is None
    stijger = next(r for r in week1 if r["Site-status"] == "stijger")
    assert isinstance(stijger["Vorige positie"], int)

    ws6 = wb["Week 06"]
    assert ws6.cell(row=2, column=1).value == "Positie"
    assert ws6.auto_filter.ref == "A2:G42"   # 7 kolommen, 40 rijen, kop op rij 2


def test_nieuwe_nummers_zijn_lichtblauw_gemarkeerd():
    """Alleen de binnenkomers krijgen kleur -- over de hele rijbreedte."""
    wb = _boek("Top40_2026.xlsx")

    ws1 = wb["Week 01"]
    koppen, week1 = _weektabel(ws1)
    # In de eerste week van onze oudste jaargang weten we niet wat er al liep,
    # dus valt de markering terug op wat de site zelf zegt. De testdata zet elke
    # zevende positie op "nieuw" (4, 11, 18, 25, 32, 39); de rest kreeg een
    # vorige positie mee en is dus geen binnenkomer.
    gemarkeerd_wk1 = _gemarkeerde_titels(ws1, koppen, week1)
    assert len(gemarkeerd_wk1) == 6, gemarkeerd_wk1
    posities_nieuw = sorted(
        r["Positie"] for r in week1 if r["Titel"] in gemarkeerd_wk1
    )
    assert posities_nieuw == [4, 11, 18, 25, 32, 39], posities_nieuw
    # En alle 40 staan er wel gewoon op -- de rest is alleen niet gekleurd.
    assert len(week1) == 40

    ws2 = wb["Week 02"]
    koppen, week2 = _weektabel(ws2)
    assert _gemarkeerde_titels(ws2, koppen, week2) == ["Liedje 39"]

    ws3 = wb["Week 03"]
    koppen, week3 = _weektabel(ws3)
    assert _gemarkeerde_titels(ws3, koppen, week3) == [ANTOON_TITEL]
    antoon = next(r for r in week3 if r["Titel"] == ANTOON_TITEL)
    assert (antoon["Positie"], antoon["Artiest"]) == (28, "Antoon ft. Sef")
    assert antoon["Site-status"] == "nieuw"        # ook volgens de site binnenkomer
    assert antoon["Sleutel"] == sleutel_van(*ANTOON_FT)

    ws4 = wb["Week 04"]
    koppen, week4 = _weektabel(ws4)
    assert _gemarkeerde_titels(ws4, koppen, week4) == ["Liedje 40"]

    # Week 6: Bløf keert terug (site-status "terug") maar is voor ons niet nieuw.
    # De rij staat er dus wel, maar zonder kleur.
    ws6 = wb["Week 06"]
    koppen, week6 = _weektabel(ws6)
    assert _gemarkeerde_titels(ws6, koppen, week6) == []
    bloef = next(r for r in week6 if r["Sleutel"] == sleutel_van(*BLOF))
    assert bloef["Site-status"] == "terug"

    ws7 = wb["Week 07"]
    koppen, week7 = _weektabel(ws7)
    assert _gemarkeerde_titels(ws7, koppen, week7) == ["Liedje 41"]


def _datum(waarde) -> date:
    """openpyxl leest een datumcel terug als datetime; wij vergelijken op de dag."""
    return waarde.date() if isinstance(waarde, datetime) else waarde


def test_totaal_punten_handmatig_nagerekend():
    wb = _boek("Top40_2026.xlsx")
    koppen, rijen = _tabel(wb["Totaal"], 1)
    assert koppen == [
        "Artiest", "Titel", "Punten", "Hoogste positie", "Aantal weken genoteerd",
        "Weken volgens site", "Binnenkomst", "Laatste notering",
        "Loopt over jaargrens", "Sleutel",
    ]
    per_sleutel = {r["Sleutel"]: r for r in rijen}

    # De lijst is elke week 40 lang, dus punten = 40 - positie + 1.
    #
    # "Nummer Eén" staat alle zes de weken op #1:
    #   6 x (40 - 1 + 1) = 6 x 40 = 240
    hit = per_sleutel[sleutel_van(*HIT)]
    assert hit["Punten"] == 240
    assert hit["Aantal weken genoteerd"] == 6
    # Volgens de site noteerde het nummer al 19 weken voor onze eerste week:
    # 19 + 6 = 25. Daarom telt onze eigen kolom 6 en de site-kolom 25.
    assert hit["Weken volgens site"] == 25
    # Week 1 van 2026 werd uitgezonden op vrijdag 2 januari (de eerste zaterdag
    # van 2026 is de 3e), week 7 zes weken later. Er staat niets van 2025 in de
    # database, dus geen enkele reeks loopt over de jaargrens.
    assert (_datum(hit["Binnenkomst"]), _datum(hit["Laatste notering"])) == (
        date(2026, 1, 2), date(2026, 2, 13))
    assert hit["Loopt over jaargrens"] is None
    assert rijen[0]["Sleutel"] == hit["Sleutel"]  # hoogste puntentotaal bovenaan

    # "Terug Van Weggeweest": week 1 #12, week 2 #19, week 3 #31, week 4 er niet
    # in, week 5 bestaat niet, week 6 #38, week 7 #35.
    #   (40-12+1) + (40-19+1) + (40-31+1) + (40-38+1) + (40-35+1)
    # =    29     +    22     +    10     +     3     +     6      = 70
    blof = per_sleutel[sleutel_van(*BLOF)]
    assert blof["Punten"] == 70
    assert blof["Aantal weken genoteerd"] == 5   # zelf geteld, niet van de site
    assert blof["Hoogste positie"] == 12
    assert (_datum(blof["Binnenkomst"]), _datum(blof["Laatste notering"])) == (
        date(2026, 1, 2), date(2026, 2, 13))

    # "Zomer In De Stad (Radio Edit)": week 3 #28, week 4 #22, week 6 #15, week 7 #11.
    #   (40-28+1) + (40-22+1) + (40-15+1) + (40-11+1)
    # =    13     +    19     +    26     +    30      = 88
    antoon = per_sleutel[sleutel_van(*ANTOON_FT)]
    assert antoon["Punten"] == 88
    assert antoon["Aantal weken genoteerd"] == 4
    assert _datum(antoon["Binnenkomst"]) == date(2026, 1, 16)   # week 3
    # Schrijfwijze veranderde van "ft." naar "feat." bij gelijke sleutel:
    # het overzicht toont de meest recente schrijfwijze.
    assert sleutel_van(*ANTOON_FT) == sleutel_van(*ANTOON_FEAT)
    assert antoon["Artiest"] == "Antoon feat. Sef"

    # Aflopend op punten gesorteerd.
    punten = [r["Punten"] for r in rijen]
    assert punten == sorted(punten, reverse=True)
    # 41 vulnummers + 3 specials = 44 unieke nummers dit jaar.
    assert len(rijen) == 44


def test_jaarmatrix_met_re_entry_en_gat():
    wb = _boek("Top40_Jaar_2026.xlsx")
    ws = wb["Jaaroverzicht"]
    koppen, rijen = _tabel(ws, 1)
    # Week 5 zit niet in de database en krijgt dus ook geen kolom.
    assert koppen == [
        "Artiest", "Titel", "1", "2", "3", "4", "6", "7",
        "Hoogste positie", "Aantal weken", "Punten", "Sleutel",
    ], koppen

    per_sleutel = {r["Sleutel"]: r for r in rijen}
    blof = per_sleutel[sleutel_van(*BLOF)]
    assert blof["1"] == 12
    assert blof["2"] == 19
    assert blof["3"] == 31
    assert blof["4"] is None          # eruit gevallen
    assert blof["6"] == 38            # re-entry
    assert blof["7"] == 35
    assert blof["Aantal weken"] == 5
    assert blof["Hoogste positie"] == 12
    assert blof["Punten"] == 70

    antoon = per_sleutel[sleutel_van(*ANTOON_FT)]
    assert [antoon[str(w)] for w in WEKEN] == [None, None, 28, 22, 15, 11]

    # Sortering: hoogste positie oplopend, daarna eerste week.
    posities = [r["Hoogste positie"] for r in rijen]
    assert posities == sorted(posities)
    assert rijen[0]["Sleutel"] == sleutel_van(*HIT)

    # Top-3-accent: posities 1 t/m 3 gekleurd, de rest niet.
    kolom_van_kop = {kop: i + 1 for i, kop in enumerate(koppen)}
    cel_een = ws.cell(row=2, column=kolom_van_kop["1"])
    assert cel_een.value == 1 and cel_een.fill.fill_type == "solid"
    for rij_index, rij in enumerate(rijen):
        for week in WEKEN:
            cel = ws.cell(row=rij_index + 2, column=kolom_van_kop[str(week)])
            gekleurd = cel.fill.fill_type == "solid"
            assert gekleurd == (cel.value in (1, 2, 3)), (rij["Titel"], week, cel.value)


def test_tipparade_wisselende_lengte():
    wb = _boek("Tipparade_2026.xlsx")
    _, totaal = _tabel(wb["Totaal"], 1)
    per_sleutel = {r["Sleutel"]: r for r in totaal}

    # De Tipparade is niet elke week even lang; punten gaan per week met de
    # lengte van DIE week (N - positie + 1):
    #   week 1: N=8,  #1 ->  8
    #   week 2: N=10, #1 -> 10
    #   week 3: N=10, #2 ->  9
    #   week 4: N=12, #1 -> 12
    #   week 6: N=9,  #3 ->  7
    #   week 7: N=11, #5 ->  7
    #   totaal: 8 + 10 + 9 + 12 + 7 + 7 = 53
    tip = per_sleutel[sleutel_van(*TIPHIT)]
    assert tip["Punten"] == 53
    assert tip["Aantal weken genoteerd"] == 6
    assert tip["Hoogste positie"] == 1

    # De weektabs volgen de wisselende lengte van de lijst.
    lengtes = {}
    for week in (1, 2, 3, 4, 6, 7):
        _, rijen = _weektabel(wb[f"Week {week:02d}"])
        lengtes[week] = len(rijen)
    assert lengtes == {1: 8, 2: 10, 3: 10, 4: 12, 6: 9, 7: 11}

    # Week 2 groeit van 8 naar 10 noteringen: twee nieuwkomers, dus twee
    # gekleurde rijen tussen de tien.
    ws2 = wb["Week 02"]
    koppen, week2 = _weektabel(ws2)
    assert sorted(_gemarkeerde_titels(ws2, koppen, week2)) == [
        "Tip Liedje 08", "Tip Liedje 09",
    ]
    ws4 = wb["Week 04"]
    koppen, week4 = _weektabel(ws4)
    assert sorted(_gemarkeerde_titels(ws4, koppen, week4)) == [
        "Tip Liedje 10", "Tip Liedje 11",
    ]
    ws6 = wb["Week 06"]
    koppen, week6 = _weektabel(ws6)
    assert len(week6) == 9                                   # lijst krimpt
    assert _gemarkeerde_titels(ws6, koppen, week6) == []     # maar niets nieuws


def test_oranje_heeft_label_en_meest_recente_schrijfwijze():
    wb = _boek("OranjeTop30_2026.xlsx")
    koppen, week1 = _weektabel(wb["Week 01"])
    assert koppen == [
        "Positie", "Vorige positie", "Artiest", "Titel", "Label",
        "Aantal weken", "Site-status", "Sleutel",
    ]
    assert len(week1) == 30
    assert week1[2]["Label"] == "Berk Music"     # positie 3 in week 1

    koppen, totaal = _tabel(wb["Totaal"], 1)
    assert koppen[:3] == ["Artiest", "Titel", "Label"]
    per_sleutel = {r["Sleutel"]: r for r in totaal}
    oranje = per_sleutel[sleutel_van(*ORANJEHIT)]
    # Label wijzigde in week 4 van "Berk Music" naar "Berk Music BV".
    assert oranje["Label"] == "Berk Music BV"
    # 30 lang, elke week op #3: 6 x (30 - 3 + 1) = 6 x 28 = 168
    assert oranje["Punten"] == 168

    jaar = load_workbook(gebouwd()["map"] / "OranjeTop30_Jaar_2026.xlsx")["Jaaroverzicht"]
    koppen, rijen = _tabel(jaar, 1)
    assert koppen[:3] == ["Artiest", "Titel", "Label"]
    assert koppen[3:9] == ["1", "2", "3", "4", "6", "7"]


def test_tabnaam_wordt_opgeschoond():
    assert excel.tabnaam("Week 1: top/40 [test]") == "Week 1 top 40 test"
    assert len(excel.tabnaam("x" * 50)) == 31
    assert all(teken not in excel.tabnaam("a:b\\c/d?e*f[g]h") for teken in ":\\/?*[]")


def test_openstaand_bestand_geeft_duidelijke_fout():
    if os.name != "nt":
        return  # elders geeft dit IsADirectoryError; die vertalen we niet
    werkmap = gebouwd()["map"]
    geblokkeerd = werkmap / "Geblokkeerd_2026.xlsx"
    geblokkeerd.mkdir(exist_ok=True)  # gedraagt zich als een bezet bestand
    try:
        excel.bewaar_werkboek(Workbook(), geblokkeerd)
    except excel.BestandInGebruik as fout:
        assert "Geblokkeerd_2026.xlsx" in str(fout)
        assert "open" in str(fout).lower()
    else:
        raise AssertionError("verwachtte BestandInGebruik")


def test_lege_lijst_levert_geen_gegevens():
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.executescript(db.SCHEMA)
    assert excel.verzamel_lijst(con, "sterrennl", JAAR) is None
    assert excel.bouw_lijst(con, "sterrennl", JAAR, gebouwd()["map"]) == []
    con.close()


def main() -> int:
    tests = [waarde for naam, waarde in sorted(globals().items()) if naam.startswith("test_")]
    mislukt = 0
    for test in tests:
        try:
            test()
        except Exception as fout:  # noqa: BLE001
            mislukt += 1
            print(f"MISLUKT  {test.__name__}: {type(fout).__name__}: {fout}")
            import traceback

            traceback.print_exc()
        else:
            print(f"ok       {test.__name__}")
    print(f"\n{len(tests) - mislukt}/{len(tests)} geslaagd")
    return 1 if mislukt else 0


if __name__ == "__main__":
    raise SystemExit(main())
